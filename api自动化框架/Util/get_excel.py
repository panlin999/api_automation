import sys
sys.path.append("D:/yiq/Demo")
import openpyxl
import os
base_path = os.getcwd()  #当前路径
path = os.path.abspath(os.path.dirname(os.getcwd()))  #上级路径

class GetExcel:


    #加载excel： 需要路径，告知哪里的excel
    def load_excel(self):
        #excel = openpyxl.load_workbook(path+'/Case/testcase.xlsx')
        excel = openpyxl.load_workbook('D:/yiq/Demo/Case/testcase.xlsx')
        return excel

    #加载excel中所有sheet信息；需要给个下标，告知是哪一个sheet，默认为第一张excel
    def get_sheet_data(self, index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]  #获取excel中sheet的数
        return data

    #获取行数
    def get_rows(self):
        data = self.get_sheet_data().max_row   #max_row 表示总行数
        return data

    #获取某一行的所有内容，那么就需要传递一个行数的参数
    def get_rows_value(self, row):
        row_list = []   #每一行的数据都是一个列表；所有先定义一个空列表
        for i in self.get_sheet_data()[row]:  #通过行数 遍历循环excel的sheet中的数据
            row_list.append(i.value)    #然后添加到首先定义好的空列表中
        return row_list

    #获取某一列的所有内容；那么就需要传递一个列的参数
    def get_columns_value(self, key=None):
        column_list = []
        if key == None:     #如果列的key是空的，那么就默认获取A列数据
            key = 'A'
        column_list_data =self.get_sheet_data()[key] #通过 列  遍历循环excel的sheet中的数据
        for i in column_list_data:
            column_list.append(i.value)
        return column_list

    #获取行号，如果要指定获取哪一行，那就需要传递一个参数以便告知，此用例中 也就可以传递一个case_id
    def get_row_number(self, case_id):
        num = 1
        cols_data =self.get_columns_value()  #首先获取某一行所有内容
        for col_data in cols_data: # 遍历循环每一行的数据，如果遍历出来的某个字段 = case_id 就返回
            if case_id == col_data:
                return num
            num +=1
        return num

    #获取某一个单元格的内容； 此时就需要传递 行 和列 俩个参数
    def get_cell_value(self, rows, cols):
        data = self.get_sheet_data().cell(row=rows, column=cols).value
        return data

    #获取excel里面的所有的数据
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i+2))
        return data_list

    #写入数据；要知道写入的行，列，以及数值
    def write_excel_data(self, rows, cols, value):
        wb = self.load_excel()  #加载excel
        wr = wb.active  #获取sheet对应对象
        wr.cell(rows, cols, value)
        #wb.save(path + '/Case/testcase.xlsx')
        wb.save('D:/yiq/Demo/Case/testcase.xlsx')

excel_data = GetExcel()

if __name__ == '__main__':
    excel_data = GetExcel()
    excel_data.load_excel()